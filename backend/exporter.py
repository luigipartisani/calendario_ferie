import io
import pandas as pd
from datetime import date, timedelta
from openpyxl.styles import PatternFill
from backend.data_processor import MESI_ITALIANI

GIORNI_IT = ['Lun', 'Mar', 'Mer', 'Gio', 'Ven', 'Sab', 'Dom']

def _easter(year):
    a = year % 19
    b, c = divmod(year, 100)
    d, e = divmod(b, 4)
    f = (b + 8) // 25
    g = (b - f + 1) // 3
    h = (19 * a + b - d - g + 15) % 30
    i, k = divmod(c, 4)
    l = (32 + 2 * e + 2 * i - h - k) % 7
    m = (a + 11 * h + 22 * l) // 451
    month = (h + l - 7 * m + 114) // 31
    day = (h + l - 7 * m + 114) % 31 + 1
    return date(year, month, day)

def _italian_holidays(year):
    easter = _easter(year)
    return {
        date(year, 1, 1), date(year, 1, 6),
        easter, easter + timedelta(days=1),
        date(year, 4, 25), date(year, 5, 1), date(year, 6, 2),
        date(year, 8, 15), date(year, 11, 1),
        date(year, 12, 8), date(year, 12, 25), date(year, 12, 26),
    }

def generate_xlsx(year, grids_ferie, selected_users):
    # Genera tutti i giorni dell'anno
    all_dates = []
    d = date(year, 1, 1)
    while d <= date(year, 12, 31):
        all_dates.append(d)
        d += timedelta(days=1)

    day_labels = [f"{d.day}/{d.month}" for d in all_dates]
    all_cols = ['Area', 'Sede', 'Dipendente'] + day_labels + ['NE', 'Totale']

    # Righe di intestazione
    row_settimana    = ['', '', ''] + [f"Settimana {d.isocalendar()[1]}" for d in all_dates] + ['', '']
    row_giorno_ddmm  = ['', '', ''] + [d.strftime('%d/%m') for d in all_dates] + ['', '']
    row_giorno_ddd   = ['', '', ''] + [GIORNI_IT[d.weekday()] for d in all_dates] + ['', '']
    row_intestazione = ['Area', 'Sede', 'Dipendente'] + [''] * (len(day_labels) + 2)

    rows = [row_settimana, row_giorno_ddmm, row_giorno_ddd, row_intestazione]

    # Righe utenti
    for user in selected_users:
        user_row = ['', '', user]
        totale = 0.0
        for d in all_dates:
            month_name = MESI_ITALIANI[d.month - 1]
            grid = grids_ferie.get(month_name)
            hours = 0.0
            if grid is not None and not grid.empty and user in grid.index and d.day in grid.columns:
                hours = grid.at[user, d.day]
            user_row.append(hours if hours > 0 else '')
            totale += hours
        user_row.append('')        # NE (vuota)
        user_row.append(totale if totale > 0 else '')
        rows.append(user_row)

    df = pd.DataFrame(rows, columns=all_cols)

    output = io.BytesIO()
    with pd.ExcelWriter(output, engine='openpyxl') as writer:
        df.to_excel(writer, index=False, header=False, sheet_name='Ferie')
        ws = writer.sheets['Ferie']

        # Merge celle della riga Settimana (riga 1) per valori consecutivi uguali
        week_numbers = [d.isocalendar()[1] for d in all_dates]
        day_col_start = 4  # colonne 1-3: Area, Sede, Dipendente
        merge_start = day_col_start
        for i in range(1, len(week_numbers)):
            if week_numbers[i] != week_numbers[i - 1]:
                if merge_start < day_col_start + i - 1:
                    ws.merge_cells(start_row=1, start_column=merge_start,
                                   end_row=1, end_column=day_col_start + i - 1)
                merge_start = day_col_start + i
        last_col = day_col_start + len(week_numbers) - 1
        if merge_start < last_col:
            ws.merge_cells(start_row=1, start_column=merge_start,
                           end_row=1, end_column=last_col)

        # Colora sabati, domeniche e festivi con #FABF8F
        holidays = _italian_holidays(year)
        fill_weekend = PatternFill(start_color='FABF8F', end_color='FABF8F', fill_type='solid')
        total_rows = ws.max_row
        for i, d in enumerate(all_dates):
            if d.weekday() >= 5 or d in holidays:
                col = day_col_start + i
                for row in range(1, total_rows + 1):
                    ws.cell(row=row, column=col).fill = fill_weekend

        # Colora riga 1 (Settimana) e riga 4 (intestazione) con #95B3D7 (sovrascrive)
        fill_header = PatternFill(start_color='95B3D7', end_color='95B3D7', fill_type='solid')
        for col in range(day_col_start, last_col + 1):
            ws.cell(row=1, column=col).fill = fill_header
        for col in range(1, len(all_cols) + 1):
            ws.cell(row=4, column=col).fill = fill_header

    output.seek(0)
    return output.getvalue()
